import openpyxl
import pandas as pd
import re

# Remove unnecessary columns
wb = openpyxl.load_workbook('fantasy_football_data.xlsx')


def clean_player_info(player_info):
    return re.sub(r'\bQ\b(?!\s*QB)', '', str(player_info)).replace("IA", "").replace("IR", "").replace("View News", "")


for sheet in wb:
    # Iteration in reverse order to avoid index shift
    for col_index in range(sheet.max_column, 0, -1):
        header = sheet.cell(1, col_index).value

        if col_index == 1 and 'Rank' in str(header):
            sheet.delete_cols(col_index)
        elif col_index == 3 and 'Opp' in str(header):
            sheet.delete_cols(col_index)
        elif col_index == 12 and 'Ret TD' in str(header):
            sheet.delete_cols(col_index)
        elif col_index == 13 and 'Misc FumTD' in str(header):
            sheet.delete_cols(col_index)
        elif col_index == 14 and 'Misc 2PT' in str(header):
            sheet.delete_cols(col_index)
        elif col_index == 19 and 'Average Points TGP' in str(header):
            sheet.delete_cols(col_index)
        elif col_index == 20 and 'Average Points TAVG' in str(header):
            sheet.delete_cols(col_index)

    # Redo column names
    for row_index in range(2, sheet.max_row + 1):
        player_cell = sheet.cell(row_index, 1)
        player_cell.value = clean_player_info(player_cell.value)

    for col_index in range(1, sheet.max_column + 1):
        header = sheet.cell(1, col_index).value
        if 'Fantasy Points' in str(header):
            sheet.cell(1, col_index, 'Total Fantasy Points')

    for col_index in range(1, sheet.max_column + 1):
        header = sheet.cell(1, col_index).value
        if 'Average Points GP' in str(header):
            sheet.cell(1, col_index, 'Total Games Played')

    for col_index in range(1, sheet.max_column + 1):
        header = sheet.cell(1, col_index).value
        if 'Average Points AVG' in str(header):
            sheet.cell(1, col_index, 'Points Per Game')

    for col_index in range(1, sheet.max_column + 1):
        header = sheet.cell(1, col_index).value
        if 'Fum Lost' in str(header):
            sheet.cell(1, col_index, 'Fumbles')

wb.save('fantasy_football_data.xlsx')
wb.close()

# Column handling
excel_file_path = 'fantasy_football_data.xlsx'
xls = pd.ExcelFile(excel_file_path)

dfs = {}

for sheet_name in xls.sheet_names:
    df = pd.read_excel(excel_file_path, sheet_name=sheet_name)

    # Handling empty spaces
    df.replace('-', pd.NA, inplace=True)
    numeric_columns = ['Passing Yds', 'Passing TD', 'Passing Int', 'Rushing Yds', 'Rushing TD', 'Receiving Rec', 'Receiving Yds', 'Receiving TD', 'Fumbles', 'Total Fantasy Points', 'Total Games Played', 'Points Per Game']
    df[numeric_columns] = df[numeric_columns].apply(pd.to_numeric, errors='coerce')

    # Insert column for ranking based on PPG
    df['Rank'] = df['Points Per Game'].rank(ascending=False, method='min', na_option='top')

    # If there are ties, adjust ranks based on "Total Fantasy Points"
    ties = df.duplicated(subset=['Rank'], keep=False)
    if any(ties):
        df.loc[ties, 'Rank'] = df.groupby('Points Per Game')['Total Fantasy Points'].rank(ascending=False, method='min', na_option='top')

    # Have rank as first column in sheet
    df = df[['Rank'] + [col for col in df.columns if col != 'Rank']]
    df = df.sort_values(by=['Rank', 'Total Fantasy Points'], ascending=[True, False]).reset_index(drop=True)

    # Get position and team
    df[['Position', 'Team']] = df['Player'].str.extract(r'(\b\w{1,2}\b)\s*-\s*(.*(?:\s*M\s*)?)')
    df['Player'] = df['Player'].replace(r'(\b\w{1,2}\b)\s*-\s*(\b\w{2,4}\b)', '', regex=True).str.strip()
    df['Player'] = df['Player'].apply(lambda player: player[:-6] if player.endswith('M') else player)
    df = df[['Rank', 'Player', 'Position', 'Team'] + [col for col in df.columns if
                                                      col not in ['Rank', 'Player', 'Position', 'Team']]]

    dfs[sheet_name] = df

with pd.ExcelWriter(excel_file_path, engine='xlsxwriter') as writer:
    for sheet_name, df in dfs.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)